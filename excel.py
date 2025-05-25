import openpyxl
import os
from tqdm import tqdm
import requests
import time
import json
from typing import Optional

def post_request(data: dict, max_retries: int = 5) -> Optional[dict]:
    headers = {
        "Host": "websbor.rosstat.gov.ru",
        "Origin": "https://websbor.rosstat.gov.ru",
        "Content-Type": "application/json",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Accept": "application/json, text/plain, */*",
        "Accept-language": "ru,en;q=0.9",
        "User-Agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/132.0.0.0 Mobile Safari/537.36"
    }
    url = "https://websbor.rosstat.gov.ru/webstat/api/gs/organizations"
    retries = 0
    while retries <= max_retries:
        try:
            response = requests.post(url, json=data, headers=headers, verify=False)
            response.raise_for_status()  # Raises HTTPError for bad responses
            # If the request is successful, return the JSON content
            return response.json()
        except requests.exceptions.HTTPError as http_err:
            if response.status_code == 429:
                retries += 1
                print(f"Too many requests. Retry {retries}/{max_retries} after 10 seconds.")
                time.sleep(7.5)
                continue  # Retry the request
            else:
                print(f"HTTP error occurred: {http_err} - Status Code: {response.status_code}")
                break  # Exit the loop for other HTTP errors
        except requests.exceptions.SSLError as ssl_err:
            print(f"SSL error occurred: {ssl_err}")
            break
        except requests.exceptions.RequestException as req_err:
            print(f"Request exception occurred: {req_err}")
            break
        except Exception as err:
            print(f"An unexpected error occurred: {err}")
            break
    print("Failed to retrieve the response after multiple retries.")
    return None

def extract_okfs_code(response_json: dict) -> Optional[str]:
    try:
        # Assuming the response is a list of organizations
        for organization in response_json:
            okfs = organization.get("okfs", {})
            code = okfs.get("code")
            if code:
                return code
        print("'code' not found in any 'okfs' field.")
    except (TypeError, AttributeError) as parse_err:
        print(f"Error parsing JSON: {parse_err}")
    return None

def get_excel_file_path():
    while True:
        path = input("Enter the path to the Excel file: ").strip()
        if os.path.isfile(path):
            if path.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
                return path
            else:
                print("The file does not seem to be an Excel file. Please enter a valid Excel file path.")
        else:
            print("File does not exist. Please enter a valid path.")

def select_sheet(workbook):
    sheets = workbook.sheetnames
    print("\nAvailable Sheets:")
    for index, sheet in enumerate(sheets, start=1):
        print(f"{index}. {sheet}")
    while True:
        try:
            choice = int(input("Select a sheet by number: "))
            if 1 <= choice <= len(sheets):
                return workbook[sheets[choice - 1]]
            else:
                print(f"Please enter a number between 1 and {len(sheets)}.")
        except ValueError:
            print("Invalid input. Please enter a number.")

def get_column_number(prompt):
    while True:
        try:
            col = int(input(prompt))
            if col >= 1:
                return col
            else:
                print("Column number must be a positive integer.")
        except ValueError:
            print("Invalid input. Please enter a number.")

def get_line_number(prompt, max_line):
    while True:
        try:
            line = int(input(prompt))
            if 1 <= line <= max_line:
                return line
            else:
                print(f"Please enter a line number between 1 and {max_line}.")
        except ValueError:
            print("Invalid input. Please enter a number.")

def main():
    print("=== Excel OKFS Code Updater ===")
    
    # Step 1: Get Excel file path
    excel_path = get_excel_file_path()

    # Load the workbook
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    # Step 2: Select sheet
    sheet = select_sheet(workbook)

    max_row = sheet.max_row

    print(f"\nSelected Sheet: {sheet.title}")
    print(f"Total Rows: {max_row}")

    # Step 3: Get column numbers
    taxpayer_col = get_column_number("Enter the column number containing the taxpayer identification number: ")
    okfs_col = get_column_number("Enter the column number for writing the OKFS code: ")

    # Step 4: Get initial and last line numbers
    initial_line = get_line_number("Enter the initial line number to start reading/writing: ", max_row)
    last_line = get_line_number("Enter the last line number to read/write: ", max_row)

    if initial_line > last_line:
        print("Initial line number cannot be greater than the last line number.")
        return

    # Step 5: Process rows
    okfs_list = []
    for row in range(initial_line, last_line + 1):
        taxpayer_cell = sheet.cell(row=row, column=taxpayer_col)
        okfs_cell = sheet.cell(row=row, column=okfs_col)

        taxpayer_id = taxpayer_cell.value
        if taxpayer_id is None:
            continue
        if okfs_cell.value is not None:
            continue
           
        # Collect the pair
        okfs_list.append((row, taxpayer_id))
    
    for line_number, tax_id in tqdm(okfs_list, desc="Processing Tax IDs", unit="tax_id"):
        okfs_cell = sheet.cell(row=line_number, column=okfs_col)
        payload = {"inn":tax_id}
        response = post_request(payload)
        okfs_code = extract_okfs_code(response)
        okfs_cell.value = okfs_code
        workbook.save(excel_path)
        time.sleep(7.5)
    
    # Save the workbook
    try:
        workbook.save(excel_path)
        print(f"\nWorkbook '{excel_path}' has been updated successfully.")
    except Exception as e:
        print(f"Error saving workbook: {e}")
        return

    # Print the collected list
    print("\nCollected Pairs (Line Number, Taxpayer Identification Number):")
    for pair in okfs_list:
        print(pair)

if __name__ == "__main__":
    main()
